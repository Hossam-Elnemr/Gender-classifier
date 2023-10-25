import openpyxl
import gender_guesser.detector as gender

excel_file = input("Enter the name of the csv (excel) file: ")
wrkbk = openpyxl.load_workbook(excel_file)
sh = wrkbk.active
d = gender.Detector(case_sensitive=False)

column_number = int(input("Enter the number of the column that has the names: "))

# Loop over each row in the selected column that has the list names

for i in range(1, sh.max_row+1):
    cell_obj = sh.cell(row=i, column = column_number)

    if cell_obj.value is None or cell_obj.value == "":
        continue
    name = cell_obj.value.split()[0]
    print(name + " : " + d.get_gender(name))
